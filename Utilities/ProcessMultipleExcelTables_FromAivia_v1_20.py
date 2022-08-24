# -------- Activate virtual environment -------------------------
import os
import ctypes
import sys
from pathlib import Path
parentFolder = str(Path(__file__).parent.parent)
activate_path = parentFolder + '\\env\\Scripts\\activate_this.py'

activate_path = r'D:\PythonCode\Python_scripts\Projects\PythonVenvForAivia\env\Scripts\activate_this.py'  # TODO: test line to remove!

if os.path.exists(activate_path):
    exec(open(activate_path).read(), {'__file__': activate_path})
    print(f'Aivia virtual environment activated\nUsing python: {activate_path}')
else:
    # Attempt to still run the script with main Aivia python interpreter
    error_mess = f'Error: {activate_path} was not found.\nPlease run the \'FirstTimeSetup.py\' script in Aivia first.'
    ans = ctypes.windll.user32.MessageBoxW(0, error_mess, 'Error', 1)
    if ans == 2:
        sys.exit(error_mess)
    print('\n'.join(['#' * 40, error_mess,
                     'Now trying to fallback on python environment specified in Aivia options > Advanced.',
                     '#' * 40]))
# ---------------------------------------------------------------

import pandas as pd
import wx
import concurrent.futures
import openpyxl.utils.cell

# Folder to quickly run the script on all Excel files in it
DEFAULT_FOLDER = ""   # "D:\\PythonCode\\Python_scripts\\Projects\\ExcelFileHandling\\tests"
DEFAULT_FOLDER = r"D:\Aivia Working Directory\_Customers\BE\Algist\22-07-04 For Training\Batch\2022-08-13-01-19-57 Workflow Living and Dead"

# Default action when combining multiple spreadsheets (see difference in scenario A and B below). False = A, True = B
do_multiple_files_as_cols = False

# Combining measurement tabs into one (for the same object subset) - For scenario A and D only
do_combine_meas_tabs = True

# Scenario D: tables processed separately
do_separated_processing = True
if do_separated_processing:
    do_multiple_files_as_cols = False
    do_combine_meas_tabs = True

"""
Convert multiple Excel spreadsheets (in the same input folder) exported from Aivia into a single Excel file.
Options:
    A - Multiple files selected, spreadsheets do not contain timepoints, data is combined in the same column (stacked data)
    B - Multiple files selected, spreadsheets do not contain timepoints, data is combined in multiple columns 
        (1 column = data from 1 spreadsheet)
    C - Multiple files selected, spreadsheets do CONTAIN timepoints, data is combined in the same column
        (1 column = 1 timepoint)
    D - List of files processed as single files, spreadsheets do not contain timepoints, 
        measurement tabs are combined as multiple columns (1 column = 1 measurement, 1 tab = 1 object subset)

WARNING: This currently works only under the following conditions:
    - The file were exported from Aivia as Excel files (not CSV)
    - There is no time dimension
    - The default row/column ordering was not changed at export.
    - Filenames do not contain '.' characters

The converted file will be saved with the same name as the original but with
"..._grouped" appended to the end.

Requirements
------------
pandas      (comes with Aivia installer)
openpyxl    (comes with Aivia installer)
xlrd        (comes with Aivia installer)
wxPython

(openpyxl and xlrd are Pandas requirements, but are not always
installed with it. Install them explicitly if you receive errors.)

Parameters
----------
aivia_excel_file : string
    Path to the Excel file exported from Aivia.

Returns
-------
DataFrame  
    Data from the spreadsheet converted to a Pandas DataFrame.

"""


# [INPUT Name:inputPath Type:string DisplayName:'Any channel']
# [OUTPUT Name:resultPath Type:string DisplayName:'Dummy to delete']
def run(params):
    global do_combine_meas_tabs, do_multiple_files_as_cols, do_separated_processing
    add_summary = False     # Used to know if the tab is missing from the beginning
    contains_tps = False    # If tables contain timepoints

    # Choose files (or rely on an hard coded default folder)
    input_folder = DEFAULT_FOLDER
    if input_folder != "":
        # Preparing file list
        all_files = os.listdir(input_folder)
        indiv_list = [os.path.join(os.path.abspath(input_folder), f) for f in all_files
                      if (f.endswith('.xlsx') and not f.endswith('_grouped.xlsx') and not f.startswith('~')
                          and f != 'Main Summary.xlsx')]

    else:
        indiv_list = pick_files()
        input_folder = os.path.dirname(indiv_list[0])

    if len(indiv_list) < 1:
        error_msg = 'No Excel file found in the selected folder:\n{}\nTry to select another folder'.format(input_folder)
        Mbox('Error', error_msg, 0)
        sys.exit(error_msg)

    # Prompt for user to see how many tables will be processed
    mess = '{} Excel files were detected.\nPress OK to continue.'.format(len(indiv_list)) + \
           '\nA confirmation popup message will inform you when the process is complete.'
    print(mess)  # for log
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(Mbox, 'Detected tables', mess, 1)
        ans = future.result()

    if ans == 2:
        sys.exit('Process terminated by user')

    # Starting point for scenario D for multiple files
    final_input_list = None
    if do_separated_processing:
        final_input_list = indiv_list                   # To prepare loop to process files independently
    else:
        final_input_list = [indiv_list[0]]                # Dummy entry which is 1 item to run the process once

    # OUTPUT folder
    output_folder = os.path.abspath(input_folder)

    # Init for a summary table (scenario D)
    df_big_summary = None

    # Main LOOP -----------------------------------------------------------------------------------------------
    for input_file in final_input_list:
        if len(final_input_list) > 1:
            indiv_list = [input_file]         # List of tables is trimmed down to one item to invoke scenario D

        # Reading first file to collect info
        df_raw_1 = pd.read_excel(indiv_list[0], sheet_name=None)

        # Check if timepoints exist (would not allow combining tabs as different columns)
        all_tabs = list(df_raw_1.keys())
        last_tab = all_tabs[-1]
        if df_raw_1[last_tab].shape[1] > 2:  # Checking no of columns in last sheet as 1st might be summary
            contains_tps = True

        # Check if summary tab is present or not
        if not any('Summary' in y for y in all_tabs):
            add_summary = True

        # defining output name
        output_basename = '{}_grouped.xlsx'.format(os.path.basename(indiv_list[0]).split('.')[0])
        output_file = os.path.join(output_folder, output_basename)

        df_grouped = {}     # init

        if len(indiv_list) == 1:    # D (see docstring)
            df_grouped = df_raw_1

        else:    # A-B-C (see docstring)
            # Collect tab names from first file
            tab_names_ref = df_raw_1.keys()

            # Processing the first file
            df_grouped = df_raw_1

            if do_multiple_files_as_cols and not contains_tps:    # B
                do_combine_meas_tabs = False     # not possible as columns = measurements

                # Renaming column headers for first table
                for t in tab_names_ref:
                    df_grouped[t].rename(columns={df_grouped[t].columns[-1]: os.path.basename(indiv_list[0])}, inplace=True)

                # Loop
                for f in indiv_list[1:]:
                    df_raw = pd.read_excel(f, sheet_name=None)
                    tab_names = df_raw.keys()

                    if tab_names == tab_names_ref:
                        # Start looping over the different sheets
                        for t in tab_names:
                            df_grouped[t] = pd.concat([df_grouped[t], df_raw[t].iloc[:, 1]], axis=1)
                            df_grouped[t].rename(columns={df_grouped[t].columns[-1]: os.path.basename(f)}, inplace=True)

            else:   # A-C
                # Adding prefix (file name) to first column
                for t in tab_names_ref:
                    df_grouped[t].iloc[:, 0] = [os.path.basename(indiv_list[0]) + "_" + r for r in df_grouped[t].iloc[:, 0]]

                # Loop
                for f in indiv_list[1:]:
                    if indiv_list != output_basename:       # avoids including an existing grouped table
                        df_raw = pd.read_excel(f, sheet_name=None)
                        tab_names = df_raw.keys()

                        if tab_names == tab_names_ref:
                            # Start looping over the different sheets
                            for t in tab_names:
                                # Adding prefix (file name) to first column in the raw table
                                df_raw[t].iloc[:, 0] = [os.path.basename(f) + "_" + r for r in df_raw[t].iloc[:, 0]]

                                # Merging to previous grouped data
                                df_grouped[t] = pd.concat([df_grouped[t], df_raw[t]], axis=0)

        # Combine tabs into one if no timepoints in data (scenario A or D)
        if not contains_tps and do_combine_meas_tabs:
            df_grouped = combine_tabs(df_grouped)

            emtpy_row = {'Summary': [''], 'Frame 0': ['']}

            # Calculate object counts
            df_summary_to_add = pd.DataFrame(emtpy_row)
            total_counts = [0] * (len(df_grouped.keys()) - 1)       # -1 for summary
            t = 0
            grand_total = 0

            for k in df_grouped.keys():
                if not k.endswith('Summary'):
                    total_counts[t] = df_grouped[k].shape[0]
                    grand_total += total_counts[t]
                    new_row = {'Summary': 'Total number_{}'.format(k), 'Frame 0': total_counts[t]}
                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                    # Report class group counts if existing
                    if 'Class Group' in df_grouped[k].columns:
                        class_group_col = df_grouped[k]['Class Group']
                        no_groups = class_group_col.max()
                        if no_groups > 1:
                            group_count = [0] * no_groups
                            for g in range(1, no_groups+1):
                                group_count[g-1] = class_group_col[class_group_col == g].count()
                                new_row = {'Summary': 'Total number_{}_Class {}'.format(k, g), 'Frame 0': group_count[g-1]}
                                df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)
                            for g in range(1, no_groups + 1):
                                percent = '{:.1%}'.format(group_count[g-1] / total_counts[t])
                                new_row = {'Summary': '{}_% of Class {}'.format(k, g), 'Frame 0': percent}
                                df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Add an empty row after each object set if classes exists
                        df_summary_to_add = df_summary_to_add.append(emtpy_row, ignore_index=True)

                    t += 1

            # Adding percentages of objects if multiple object sets exists
            if len(total_counts) > 1:
                for t in range(len(total_counts)):
                    val = '{:.1%}'.format(total_counts[t] / grand_total)
                    new_row = {'Summary': '% of {}'.format([*df_grouped][t+1]), 'Frame 0': val}
                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

            # Add the summary tab
            if add_summary:
                df_summary = pd.DataFrame(emtpy_row)
                df_grouped['Summary'] = df_summary

            # Merge with potential existing summary tab
            df_grouped['Summary'] = pd.concat([df_grouped['Summary'], df_summary_to_add], axis=0, ignore_index=True)

        # Writing sheets to excel
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Write Summary first
            df_grouped['Summary'].to_excel(writer, sheet_name='Summary', index=False)

            # Resizing columns
            for c in range(0, len(df_grouped['Summary'].columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = df_grouped['Summary'].iloc[:, c].map(str).str.len().max()
                writer.sheets['Summary'].column_dimensions[col_letter].width = len_longest_text

            for sh in [d for d in df_grouped.keys() if d != 'Summary']:
                df_grouped[sh].to_excel(writer, sheet_name=sh, index=False)

                # Resizing columns
                for c in range(0, len(df_grouped[sh].columns)):
                    col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                    len_longest_text = len(str(df_grouped[sh].columns[c]))
                    if c == 0:
                        if df_grouped[sh].iloc[0, 1] > 1:
                            len_longest_text = len(str(df_grouped[sh].iloc[0, 1]))
                    writer.sheets[sh].column_dimensions[col_letter].width = len_longest_text

        # Create a final summary file if multiple tables were saved
        if len(final_input_list) > 1:
            filename = os.path.basename(indiv_list[0]).split('.')[0]
            if df_big_summary is None:
                df_big_summary = df_grouped['Summary']

                # Rename header of 2nd column
                df_big_summary.rename(columns={df_big_summary.columns[1]: filename}, inplace=True)

            else:
                df_big_summary[filename] = df_grouped['Summary'][df_grouped['Summary'].columns[1]]

    # Main LOOP -----------------------------------------------------------------------------------------------

    final_mess = '{} tables were saved here:\n{}'.format(len(final_input_list), output_folder)

    # Write final summary file if multiple tables were saved
    if len(final_input_list) > 1 and not contains_tps:
        # defining output name
        output_basename = 'Main Summary'
        output_file = os.path.join(output_folder, output_basename + '.xlsx')

        # Write file
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_big_summary.to_excel(writer, sheet_name=output_basename, index=False)

            # Resizing columns
            for c in range(0, len(df_big_summary.columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = max([df_big_summary.iloc[:, c].map(str).str.len().max(), len(df_big_summary.columns[c])])
                writer.sheets[output_basename].column_dimensions[col_letter].width = len_longest_text

        final_mess += f'\n\nA main summary table was also saved as \'{output_basename}.xlsx\'.'

    # Message box to confirm table processing
    print(final_mess)
    Mbox('Table processed', final_mess, 0)


def combine_tabs(df_raw):
    df_combined = {}
    df_temp = pd.DataFrame()
    object_name = ''
    summary_exists = False

    for k in df_raw.keys():
        # Don't need the summary tab if included
        if k == 'Summary' or '.Summary' in k:
            if not summary_exists:
                df_combined = {'Summary': df_raw[k]}
                summary_exists = True

                # Add header as row
                df_combined['Summary'] = pd.DataFrame([['', ''], df_combined['Summary'].columns.values.tolist()],
                                                      columns=df_combined['Summary'].columns).append(df_combined['Summary'])
                # Rename 1st column name to standardize it
                df_combined['Summary'].rename(columns={df_combined['Summary'].columns[0]: 'Summary'}, inplace=True)

            else:           # If there is a 2nd summary tab (2nd object set)
                # Add header as row
                df_sum_temp = pd.DataFrame([['', ''], df_raw[k].columns.values.tolist()], columns=df_raw[k].columns).append(df_raw[k])

                df_sum_temp.columns = df_combined['Summary'].columns
                df_combined['Summary'] = pd.concat([df_combined['Summary'], df_sum_temp], axis=0, ignore_index=True)

        # First iteration with detailed objects
        elif k != 'Summary' and df_temp.empty is True:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = df_raw[k].columns[0]

            # Copying the sheet
            df_temp = df_raw[k]

            # Writing headers for the 1st and 2nd column
            df_temp.columns = [object_name, meas_name]

        # Fill the dataframe
        else:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name_temp = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = (df_raw[k].columns[0])[len(object_name)+1:]

            # Check if object name changed or not
            if object_name_temp != object_name and object_name_temp != '':
                # Adding prepared sheet to main series to create a new sheet
                df_combined[object_name] = df_temp

                # Now using new name as the new reference
                object_name = object_name_temp

                # Copying the current read sheet to be a new one
                df_temp = df_raw[k]

                # Writing headers for the 1st and 2nd column
                df_temp.columns = [object_name, meas_name]

            else:
                # Adding the new column to existing temp sheet
                df_temp = pd.concat([df_temp, df_raw[k].iloc[:, 1]], axis=1)

                # Adding the measurement name as a header
                df_temp.rename(columns={df_temp.columns[-1]: meas_name}, inplace=True)

    # Adding a generic name to objects if none
    if object_name == '':
        object_name = 'Object Set 1'

    # Adding last prepared sheet to main series to create a new sheet
    df_combined[object_name] = df_temp

    return df_combined


def get_split_name(txt: str):
    # First check if text doesn't end with ...
    if txt.endswith('...'):
        txt = txt[:-3]
        meas_name = '--incomplete--'        # name can't be retrieved from here
    else:
        meas_name = txt.split('.')[-1]

    obj_name = '.'.join(txt.split('.')[:-1])
    if obj_name == 'Std. Dev':
        meas_name = txt
        obj_name = ''

    return meas_name, obj_name


def pick_files():
    print('Starting wxPython app')
    app = wx.App()

    # Create open file dialog
    openFileDialog = wx.FileDialog(None, "Select a results table (xlsx) to process", ".\\", "",
                                   "Excel files (*.xlsx)|*.xlsx", wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)

    openFileDialog.ShowModal()
    filenames = openFileDialog.GetPaths()
    print("Selected table(s): ", filenames)
    openFileDialog.Destroy()
    return filenames


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


if __name__ == '__main__':
    params = {}
    run(params)

# Changelog:
# v1.00: - using wxPython for the file picker, multiple sheets stored as dictionary of DataFrames (keys = sheet names)
# v1.01: - Adding default values before def run
# v1.02: - Exception handling when measurement name contains dots: 'Std. Dev. Intensity'
# v1.10: - Adding summary tab if needed with object set count.
#        - Handles multiple summary tabs and adds counting and % values
#        - Added virtual environment activation
# v1.20: - Renaming script to "Process..."
#        - Scenario D (individual process) becomes available for multiple selected files
#        - Adding a summary table on top of individual files for scenario D
