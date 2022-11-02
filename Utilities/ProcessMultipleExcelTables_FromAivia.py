# -------- Activate virtual environment -------------------------
import os
import ctypes
import sys
from pathlib import Path

parentFolder = str(Path(__file__).parent.parent)
activate_path = parentFolder + '\\env\\Scripts\\activate_this.py'

if os.path.exists(activate_path):
    exec(open(activate_path).read(), {'__file__': activate_path})
    print(f'Aivia virtual environment activated\nUsing python: {activate_path}')
else:
    # Attempt to still run the script with main Aivia python interpreter
    error_mess = f'Error: {activate_path} was not found.\nPlease run the \'FirstTimeSetup.py\' script in Aivia first.'
    ans = ctypes.windll.user32.MessageBoxW(0, error_mess, 'Error', 1)
    if ans == 2:
        sys.exit(error_mess)
    print('\n'.join(['#' * 40, error_mess,
                     'Now trying to fallback on python environment specified in Aivia options > Advanced.',
                     '#' * 40]))
# ---------------------------------------------------------------

import pandas as pd
import wx
import concurrent.futures
import openpyxl.utils.cell
from pathlib import Path
from magicgui import magicgui
import re
from datetime import datetime

# Folder to quickly run the script on all Excel files in it
DEFAULT_FOLDER = ''

# Collect scenario
scenario_descriptions = ['A: Select multiple xlsx tables to create a combined table.\n'
                         'Same measurements are combined in the same column '
                         '(stacked data, 1 column = 1 measurement type).\n'
                         'Not compatible with timelapses.',
                         'B: Select multiple xlsx tables to create a combined table.\n'
                         'Measurements are combined as multiple columns (1 column = data from 1 xlsx table).\n',
                         'C: Select multiple xlsx tables to create a combined table.\n'
                         'Compatibility with timelapses expected (1 column = 1 timepoint).',
                         'D: Select one OR multiple xlsx tables to be processed individually.\n'
                         'Measurement tabs are combined as multiple columns '
                         '(1 column = 1 measurement, 1 tab = 1 object subset)\n'
                         'Not compatible with timelapses.',
                         'E: [From Workflow / Aivia 11.0+] Select one xlsx table, '
                         'an automatic search is performed to process other tables in the same batch.\n'
                         'One table leads to one new table where data is combined as 1 column = 1 measurement.\n'
                         'Also creates an "Analysis Summary" xlsx table, combining all "Summary" tabs from all files.\n'
                         'If a multiwell format exists, the "Analysis Summary" will group images per well.\n'
                         'Not compatible with timelapses.',
                         'F: [From Workflow / Aivia 11.0+] Select one xlsx table, '
                         'an automatic search is performed to process other '
                         'tables in the same batch.\n'
                         'Creates only an "Analysis_All results" xlsx table, combining all values '
                         '(1 column = 1 measurement from 1 table) from all files.\n'
                         'If a multiwell format exists, data of images in the same well are stacked altogether '
                         '(becomes 1 column = 1 measurement from 1 well).\n'
                         'Not compatible with timelapses.'
                         ]


@magicgui(scenario={"label": "Select a scenario:", "widget_type": "RadioButtons", 'choices': scenario_descriptions},
          call_button="Run")
def get_scenario(scenario=scenario_descriptions[5]):
    pass


@get_scenario.called.connect
def close_GUI_callback():
    get_scenario.close()


get_scenario.show(run=True)
selected_scenario = get_scenario.scenario.value[0]

"""
Convert multiple Excel spreadsheets (in the same input folder) exported from Aivia into a single Excel file.
Options:
    A - Multiple files selected, spreadsheets do not contain timepoints, data is combined in the same column (stacked data)
    B - Multiple files selected, spreadsheets do not contain timepoints, data is combined in multiple columns 
        (1 column = data from 1 spreadsheet)
    C - Multiple files selected, spreadsheets do CONTAIN timepoints, data is combined in the same column
        (1 column = 1 timepoint)
    D - List of files processed as single files, spreadsheets do not contain timepoints, 
        measurement tabs are combined as multiple columns (1 column = 1 measurement, 1 tab = 1 object subset)
    E - Same as D but only one file is selected. Expecting Aivia 11.0 subfolder hierarchy from a multiwell Workflow
        batch processing 
    ... see other info above in scenario descriptions

WARNING: This currently works only under the following conditions:
    - The file were exported from Aivia as Excel files (not CSV)
    - There is no time dimension
    - The default row/column ordering was not changed at export.
    - Filenames do not contain '.' characters

The converted file will be saved with the same name as the original but with
"..._grouped" appended to the end.

Requirements
------------
pandas
openpyxl
xlrd
wxPython
magicgui

Parameters
----------
aivia_excel_file : string
    Path to the Excel file exported from Aivia.

Returns
-------
DataFrame  
    Data from the spreadsheet converted to a Pandas DataFrame.

"""


# [INPUT Name:inputPath Type:string DisplayName:'Any channel']
# [OUTPUT Name:resultPath Type:string DisplayName:'Dummy to delete']
def run(params):
    global selected_scenario
    do_multiple_files_as_cols = False  # Default action when combining multiple spreadsheets - For B only
    do_combine_meas_tabs = False  # Combining measurement tabs into one (for the same object subset) For scenario A and D only
    do_separated_processing = False  # If tables are processed separately - only True for D and E
    do_scan_workflow_folders = False  # Only for E

    # Defining actions depending on scenario type
    if selected_scenario == 'A':
        do_combine_meas_tabs = True

    elif selected_scenario == 'B':
        do_multiple_files_as_cols = True

    elif selected_scenario == 'D':
        do_combine_meas_tabs = True
        do_separated_processing = True

    elif selected_scenario == 'E':
        do_combine_meas_tabs = True
        do_separated_processing = True
        do_scan_workflow_folders = True

    elif selected_scenario == 'F':
        do_multiple_files_as_cols = True
        do_scan_workflow_folders = True

    add_summary = False  # Used to know if the tab is missing from the beginning
    contains_tps = False  # If tables contain timepoints

    # Choose files (or rely on an hard coded default folder)
    input_folder = DEFAULT_FOLDER
    if input_folder != "":
        # Preparing file list
        all_files = os.listdir(input_folder)
        indiv_path_list = [os.path.join(os.path.abspath(input_folder), f) for f in all_files
                           if (f.endswith('.xlsx') and not f.endswith('_grouped.xlsx') and not f.startswith('~')
                               and f != 'Analysis Summary.xlsx')]

    else:
        indiv_path_list = pick_files()
        input_folder = os.path.dirname(indiv_path_list[0])

    # Scenario E-F: Collecting main folder
    batch_path = ''
    well_ref_for_tables = []  # Expected structure = batch \ A1 \ Job 1 \ Measurements \
    if do_scan_workflow_folders:
        if os.path.basename(input_folder) != 'Measurements':
            error_msg = 'This folder ({}) is expected to be named "Measurements".' \
                        '\nSelect a table again...'.format(input_folder)
            Mbox('Error', error_msg, 0)
            sys.exit(error_msg)

        indiv_path_list = []
        batch_path = str(Path(input_folder).parents[2])  # 3 levels up

        # Search files in subfolders
        for well_f in [os.path.join(batch_path, wf) for wf in os.listdir(batch_path)
                       if os.path.isdir(os.path.join(batch_path, wf))]:  # well level
            for fov_f in [os.path.join(well_f, fov) for fov in os.listdir(well_f)
                          if os.path.isdir(os.path.join(well_f, fov))]:  # fov level
                for f in [x for x in os.listdir(os.path.join(fov_f, 'Measurements')) if x.endswith('.xlsx')]:
                    indiv_path_list.append(os.path.join(os.path.join(fov_f, 'Measurements'), f))
                    well_ref_for_tables.append(os.path.basename(well_f))

    if len(indiv_path_list) < 1:
        error_msg = 'No Excel file found in the selected folder:\n{}\nTry to select another folder'.format(input_folder)
        Mbox('Error', error_msg, 0)
        sys.exit(error_msg)

    # Prompt for user to see how many tables will be processed
    mess = '{} Excel files were detected.\nPress OK to continue.'.format(len(indiv_path_list)) + \
           '\nA confirmation popup message will inform you when the process is complete.'
    print(mess)  # for log
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(Mbox, 'Detected tables', mess, 1)
        ans = future.result()

    if ans == 2:
        sys.exit('Process terminated by user')

    # Starting point for scenario D for multiple files
    if do_separated_processing:
        final_input_list = indiv_path_list  # To prepare loop to process files independently
    else:
        final_input_list = [indiv_path_list[0]]  # Dummy entry which is 1 item to run the process once

    # OUTPUT folder
    output_folder = os.path.abspath(input_folder)
    if do_scan_workflow_folders:
        output_folder = batch_path

    # Init for a summary table (scenario D-E-F)
    df_big_summary = pd.DataFrame()
    tab_ind = 0  # Tab index for pooling of summary tabs from the same well
    temp_tab = []  # Used to store tabs from the same well
    well_ref = ''

    # Evaluate time if number of file is > 10
    if len(final_input_list[:]) > 10:
        t1 = datetime.now()

    # Main LOOP -----------------------------------------------------------------------------------------------
    for file_index, input_file in enumerate(final_input_list):
        if len(final_input_list) > 1:
            indiv_path_list = [input_file]  # List of tables is trimmed down to one item to invoke scenario D

        # Evaluate time if number of file is > 10
        if len(indiv_path_list[:]) > 10:
            t1 = datetime.now()

        print('Processing: ', input_file)

        # Reading first file to collect info
        df_raw_1 = pd.read_excel(indiv_path_list[0], sheet_name=None)

        # Check if timepoints exist (would not allow combining tabs as different columns)
        all_tabs = list(df_raw_1.keys())
        last_tab = all_tabs[-1]
        if df_raw_1[last_tab].shape[1] > 2:  # Checking no of columns in last sheet as 1st might be summary
            contains_tps = True

        # Check if summary tab is present or not
        if not any('Summary' in y for y in all_tabs):
            add_summary = True

        # defining output name
        if do_multiple_files_as_cols:
            output_basename = 'Analysis_All results.xlsx'
        else:
            output_basename = '{}_grouped.xlsx'.format(os.path.basename(indiv_path_list[0]).split('.')[0])
        output_file = os.path.join(output_folder, output_basename)

        df_grouped = {}  # init

        if len(indiv_path_list) == 1:  # D (see docstring)
            df_grouped = df_raw_1

        else:  # A-B-C... (see docstring)
            # Collect tab names from first file
            tab_names_ref = df_raw_1.keys()

            if do_multiple_files_as_cols and not contains_tps:  # B-F
                do_combine_meas_tabs = False  # not possible as columns = measurements

                # Detect multiwell batch
                process_wells = is_multiwell(well_ref_for_tables[0])  # TODO: Remove summary tab if True?

                # First table in final table
                df_grouped = df_raw_1

                # Renaming column headers for first table
                if not process_wells:
                    for t in tab_names_ref:
                        df_grouped[t].rename(columns={df_grouped[t].columns[-1]: os.path.basename(indiv_path_list[0])},
                                             inplace=True)

                # Loop
                tab_names = ''
                first_well = True
                temp_df = {}
                print('-- {} processed (1/{}).'.format(os.path.basename(indiv_path_list[0]), len(indiv_path_list)))

                for f_index, f in enumerate(indiv_path_list[1:]):
                    df_raw = pd.read_excel(f, sheet_name=None)
                    tab_names = df_raw.keys()

                    # Check if well changes
                    same_well = (well_ref_for_tables[f_index + 1] == well_ref_for_tables[f_index])

                    if tab_names == tab_names_ref:
                        # Start looping over the different sheets
                        for t in tab_names:
                            if process_wells and same_well:
                                if first_well:
                                    df_grouped[t] = pd.concat([df_grouped[t], df_raw[t]], axis=0, ignore_index=True)
                                else:
                                    # Add data as stacked in the current (temp) column
                                    temp_df[t] = pd.concat([temp_df[t], df_raw[t]], axis=0, ignore_index=True)

                            else:  # Add data as a new column
                                # Push existing temp column to final table
                                if process_wells:
                                    if not first_well:
                                        df_grouped[t] = pd.concat([df_grouped[t], temp_df[t].iloc[:, 1]], axis=1)

                                    # Storing data in a new stacked column
                                    temp_df[t] = df_raw[t]

                                else:
                                    df_grouped[t] = pd.concat([df_grouped[t], df_raw[t].iloc[:, 1]], axis=1)

                                # Renaming column header
                                if process_wells:   # header is well name
                                    df_grouped[t].rename(
                                        columns={df_grouped[t].columns[-1]: well_ref_for_tables[f_index]}, inplace=True)
                                else:               # Header is file name
                                    df_grouped[t].rename(columns={df_grouped[t].columns[-1]: os.path.basename(f)},
                                                         inplace=True)

                        if not same_well:
                            first_well = False

                        print('-- {} processed ({}/{}).'.format(os.path.basename(f), f_index + 2, len(indiv_path_list)))

                    else:
                        print('WARNING: {} table was excluded because table tabs are different from first table'.format(f))

                    # Evaluate time for the processing of one table
                    if len(indiv_path_list[:]) > 10 and f_index == 0:
                        show_estimated_time(t1, len(indiv_path_list[:]))

                # Push the latest stacked column in the final table (in case of multiwell)
                for t in tab_names:
                    if process_wells:
                        if temp_df:
                            df_grouped[t] = pd.concat([df_grouped[t], temp_df[t].iloc[:, 1]], axis=1)
                        df_grouped[t].rename(columns={df_grouped[t].columns[-1]: well_ref_for_tables[-1]}, inplace=True)

            else:  # A-C
                # Adding prefix (file name) to first column
                for t in tab_names_ref:
                    df_grouped[t].iloc[:, 0] = [os.path.basename(indiv_path_list[0]) + "_" + r for r in
                                                df_grouped[t].iloc[:, 0]]

                # Loop
                for f in indiv_path_list[1:]:
                    if indiv_path_list != output_basename:  # avoids including an existing grouped table
                        df_raw = pd.read_excel(f, sheet_name=None)
                        tab_names = df_raw.keys()

                        if tab_names == tab_names_ref:
                            # Start looping over the different sheets
                            for t in tab_names:
                                # Adding prefix (file name) to first column in the raw table
                                df_raw[t].iloc[:, 0] = [os.path.basename(f) + "_" + r for r in df_raw[t].iloc[:, 0]]

                                # Merging to previous grouped data
                                df_grouped[t] = pd.concat([df_grouped[t], df_raw[t]], axis=0)

                        # Evaluate time for the processing of one table
                        if len(indiv_path_list[:]) > 10 and f == indiv_path_list[1]:
                            show_estimated_time(t1, len(indiv_path_list[1:]))

        # --- COMBINE TABS into one if no timepoints in data (scenario A-D...) ------------------------------
        if not contains_tps and (do_combine_meas_tabs or do_multiple_files_as_cols):
            # Init
            col_headers = ['Summary', *list(df_grouped[list(df_grouped.keys())[-1]].columns[1:])]
            empty_list = [['']] * len(col_headers)
            empty_row = dict(zip(col_headers, empty_list))

            # Calculate object counts
            df_summary_to_add = pd.DataFrame(empty_row)
            total_counts = []
            t = 0
            grand_total = 0

            # Scenario A or D ----------------------
            if do_combine_meas_tabs:
                df_grouped = combine_tabs(df_grouped)

                for k in df_grouped.keys():
                    if not k.endswith('Summary'):
                        total_counts.append(df_grouped[k].shape[0])
                        grand_total += total_counts[t]
                        new_row = dict(zip(list(empty_row.keys()), ['Total number_{}'.format(k), total_counts[t]]))
                        df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Report class group counts if existing
                        if 'Class Group' in df_grouped[k].columns:
                            class_group_col = df_grouped[k]['Class Group']
                            no_groups = class_group_col.max()
                            if no_groups > 1:
                                group_count = [0] * no_groups
                                for g in range(1, no_groups + 1):
                                    group_count[g - 1] = class_group_col[class_group_col == g].count()
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['Total number_{}_Class {}'.format(k, g), group_count[g - 1]]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)
                                for g in range(1, no_groups + 1):
                                    percent = '{:.1%}'.format(group_count[g - 1] / total_counts[t])
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['{}_% of Class {}'.format(k, g), percent]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                            # Add an empty row after each object set if classes exists
                            df_summary_to_add = df_summary_to_add.append(empty_row, ignore_index=True)

                        t += 1

            # Scenario F ---------------
            if do_multiple_files_as_cols:
                # Chasing counts only for object sets, not for single measurements
                object_set_ref = ''
                for k in df_grouped.keys():
                    _, object_set = get_split_name(k)
                    if object_set == '':        # If only one object set, name is not present
                        object_set = 'Object 1'

                    if not k.endswith('Summary') and not k.endswith('Class Group'):
                        if object_set != object_set_ref:
                            total_counts.append(df_grouped[k].count()[1:])
                            new_row = dict(zip(list(empty_row.keys()),
                                               ['Total number_{}'.format(object_set), *total_counts[t]]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                            object_set_ref = object_set
                            t += 1

                    elif k.endswith('Class Group'):
                        class_group_data = df_grouped['Class Group']
                        group_count = [[]] * len(df_grouped[k].columns - 1)
                        total_counts_class = [0] * len(df_grouped[k].columns - 1)
                        group_max = 0

                        for col_index, col_name in enumerate(class_group_data.columns):
                            col_values = class_group_data[col_name]
                            no_groups = col_values.max()
                            if no_groups > group_max:
                                group_max = no_groups
                            group_count[col_index] = [col_values[col_values == g + 1].count() for g in range(no_groups)]
                            total_counts_class[col_index] = len(col_values)

                        # Reconstitute data per row (per group) and add values
                        group_count_per_row = [0] * len(group_count)
                        class_percent_per_row = [''] * len(group_count)
                        for g in range(group_max):
                            for c in range(len(group_count)):
                                if len(group_count[c]) >= g + 1:
                                    group_count_per_row[c] = group_count[c][g]
                                    class_percent_per_row[c] = '{:.1%}'.format(
                                        group_count_per_row[c] / total_counts_class[c])
                                else:
                                    group_count_per_row[c] = 0
                                    class_percent_per_row[c] = '{:.1%}'.format(0)

                            new_row = dict(zip(list(empty_row.keys()),
                                               ['Total number_{}_Class {}'.format(object_set, g + 1),
                                                *group_count_per_row]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Add percentages
                        for g in range(group_max):
                            new_row = dict(zip(list(empty_row.keys()),
                                               ['{}_% of Class {}'.format(object_set, g + 1), *class_percent_per_row]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Add an empty row after each object set if classes exists
                        df_summary_to_add = df_summary_to_add.append(empty_row, ignore_index=True)

                        t += 1

            # Adding percentages of objects if multiple object sets exists
            if do_combine_meas_tabs:  # TODO: for do_multiple_files_as_cols too??
                if len(total_counts) > 1:
                    # Collect tab names without summary
                    df_grouped_keys_nosum = [k for k in df_grouped.keys() if k != 'Summary']
                    for t in range(len(total_counts)):
                        val = '{:.1%}'.format(total_counts[t] / grand_total)
                        new_row = {'Summary': '% of {}'.format(df_grouped_keys_nosum[t]), 'Frame 0': val}
                        df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

            # Add the summary tab
            if add_summary:
                df_summary = pd.DataFrame(empty_row)
                df_grouped['Summary'] = df_summary

            # Merge with potential existing summary tab
            df_grouped['Summary'] = pd.concat([df_grouped['Summary'], df_summary_to_add], axis=0, ignore_index=True)

            # Removing double empty rows
            df_grouped['Summary'] = remove_double_empty_rows(df_grouped['Summary'])

        # --- WRITING EXCEL FILES -----------------------------------------------------------------------------------
        # Writing sheets to excel
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Write Summary first
            df_grouped['Summary'].to_excel(writer, sheet_name='Summary', index=False)

            # Resizing columns
            for c in range(0, len(df_grouped['Summary'].columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = df_grouped['Summary'].iloc[:, c].map(str).str.len().max()
                writer.sheets['Summary'].column_dimensions[col_letter].width = len_longest_text * 1.5

            for sh in [d for d in df_grouped.keys() if d != 'Summary']:
                df_grouped[sh].to_excel(writer, sheet_name=sh, index=False)

                # Resizing columns
                for c in range(0, len(df_grouped[sh].columns)):
                    col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                    len_longest_text = len(str(df_grouped[sh].columns[c]))
                    if c == 0:  # First column with measurement name and object names
                        if len(str(df_grouped[sh].iloc[1, 0])) > len_longest_text:
                            len_longest_text = len(str(df_grouped[sh].iloc[1, 0]))
                    if len_longest_text < 10:
                        len_longest_text = 10
                    writer.sheets[sh].column_dimensions[col_letter].width = len_longest_text

        # --- FINAL SUMMARY ----------------------------------------------------------------------------------------
        # Create a final summary file if multiple tables were saved
        if len(final_input_list) > 1 and not do_multiple_files_as_cols:
            filename = os.path.basename(indiv_path_list[0]).split('.')[0]

            if do_scan_workflow_folders:
                # Checking well name compared to previous table
                current_well = well_ref_for_tables[file_index]
                if current_well != well_ref or file_index == len(final_input_list) - 1:
                    if temp_tab:
                        # Combine summary tabs and add to final super table

                        # Add latest summary tab to final super table
                        if file_index == len(final_input_list) - 1:
                            temp_tab.append(df_grouped['Summary'][df_grouped['Summary'].columns[1]])

                        if len(temp_tab) > 1:
                            # Add empty table to create space between parameters
                            temp_tab.append(pd.Series('', index=range(temp_tab[0].shape[0])))

                            # Init table
                            well_tab = temp_tab[0].iloc[[0]]

                            # Add first row of other tables
                            for tab in temp_tab[1:]:
                                well_tab = well_tab.append(pd.Series(tab.iloc[[0]]), ignore_index=True)

                            # Add other rows
                            for r in range(1, temp_tab[0].shape[0]):
                                for tab in temp_tab:
                                    well_tab = well_tab.append(pd.Series(tab.iloc[[r]]), ignore_index=True)

                            # Drop empty rows
                            if isinstance(well_tab, pd.Series):
                                well_tab = well_tab.to_frame()
                            well_tab = remove_double_empty_rows(well_tab)

                        else:
                            well_tab = temp_tab[0]

                        # Push into the big summary
                        if df_big_summary.empty:
                            df_big_summary = well_tab  # init
                            # Rename header of well column
                            df_big_summary.rename(columns={df_big_summary.columns[0]: well_ref}, inplace=True)

                        else:
                            df_big_summary[well_ref] = well_tab

                    well_ref = current_well
                    temp_tab = []

                # Add first column only if temp_tab is empty
                if not temp_tab:
                    temp_tab.append(df_grouped['Summary'][df_grouped['Summary'].columns[0]])

                # Add data to be combined
                temp_tab.append(df_grouped['Summary'][df_grouped['Summary'].columns[1]])

            else:
                if df_big_summary is None:
                    df_big_summary = df_grouped['Summary']

                    # Rename header of 2nd column
                    df_big_summary.rename(columns={df_big_summary.columns[1]: filename}, inplace=True)

                else:
                    df_big_summary[filename] = df_grouped['Summary'][df_grouped['Summary'].columns[1]]

        # --- / FINAL SUMMARY ----------------------------------------------------------------------------------------

        # Evaluate time for one table if a list of table is processed individually
        if len(final_input_list[:]) > 10 and file_index == 0:
            show_estimated_time(t1, len(final_input_list[:]))

    # Main LOOP -----------------------------------------------------------------------------------------------

    final_mess = '{} table(s) got saved here:\n{}'.format(len(final_input_list), output_folder)

    # Write final summary file if multiple tables were saved
    if len(final_input_list) > 1 and not contains_tps and not do_multiple_files_as_cols:
        # defining output name
        output_basename = 'Analysis Summary'
        output_file = os.path.join(output_folder, output_basename + '.xlsx')

        # Write file
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_big_summary.to_excel(writer, sheet_name=output_basename, index=False)

            # Resizing columns
            for c in range(0, len(df_big_summary.columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = max(
                    [df_big_summary.iloc[:, c].map(str).str.len().max(), len(df_big_summary.columns[c])])
                writer.sheets[output_basename].column_dimensions[col_letter].width = len_longest_text

        final_mess += f'\n\nA main summary table was also saved as \'{output_basename}.xlsx\'.'

    # Message box to confirm table processing
    print(final_mess)
    Mbox('Table processed', final_mess, 0)


# Combine tabs of a single spreadsheet file
def combine_tabs(df_raw):
    df_combined = {}
    df_temp = pd.DataFrame()
    object_name = ''
    summary_exists = False

    for k in df_raw.keys():
        # Don't need the summary tab if included
        if k == 'Summary' or '.Summary' in k:
            if not summary_exists:
                df_combined = {'Summary': df_raw[k]}
                summary_exists = True

                # Add header as row
                df_combined['Summary'] = pd.DataFrame([['', ''], df_combined['Summary'].columns.values.tolist()],
                                                      columns=df_combined['Summary'].columns).append(
                    df_combined['Summary'])
                # Rename 1st column name to standardize it
                df_combined['Summary'].rename(columns={df_combined['Summary'].columns[0]: 'Summary'}, inplace=True)

            else:  # If there is a 2nd summary tab (2nd object set)
                # Add header as row
                df_sum_temp = pd.DataFrame([['', ''], df_raw[k].columns.values.tolist()],
                                           columns=df_raw[k].columns).append(df_raw[k])

                df_sum_temp.columns = df_combined['Summary'].columns
                df_combined['Summary'] = pd.concat([df_combined['Summary'], df_sum_temp], axis=0, ignore_index=True)

        # First iteration with detailed objects
        elif k != 'Summary' and df_temp.empty is True:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = df_raw[k].columns[0]

            # Copying the sheet
            df_temp = df_raw[k]

            # Writing headers for the 1st and 2nd column
            df_temp.columns = [object_name, meas_name]

        # Fill the dataframe
        else:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name_temp = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = (df_raw[k].columns[0])[len(object_name) + 1:]

            # Check if object name changed or not
            if object_name_temp != object_name and object_name_temp != '':
                # Adding prepared sheet to main series to create a new sheet
                df_combined[object_name] = df_temp

                # Now using new name as the new reference
                object_name = object_name_temp

                # Copying the current read sheet to be a new one
                df_temp = df_raw[k]

                # Writing headers for the 1st and 2nd column
                df_temp.columns = [object_name, meas_name]

            else:
                # Adding the new column to existing temp sheet
                df_temp = pd.concat([df_temp, df_raw[k].iloc[:, 1]], axis=1)

                # Adding the measurement name as a header
                df_temp.rename(columns={df_temp.columns[-1]: meas_name}, inplace=True)

    # Adding a generic name to objects if none
    if object_name == '':
        object_name = 'Object Set 1'

    # Adding last prepared sheet to main series to create a new sheet
    df_combined[object_name] = df_temp

    return df_combined


def get_split_name(txt: str):
    # First check if text doesn't end with ...
    if txt.endswith('...'):
        txt = txt[:-3]
        meas_name = '--incomplete--'  # name can't be retrieved from here
    else:
        meas_name = txt.split('.')[-1]

    obj_name = '.'.join(txt.split('.')[:-1])
    if obj_name == 'Std. Dev':
        meas_name = txt
        obj_name = ''

    return meas_name, obj_name


def pick_files():
    print('Starting wxPython app')
    app = wx.App()

    # Create open file dialog
    openFileDialog = wx.FileDialog(None, "Select a results table (xlsx) to process", ".\\", "",
                                   "Excel files (*.xlsx)|*.xlsx", wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)

    openFileDialog.ShowModal()
    filenames = openFileDialog.GetPaths()
    print("Selected table(s): ", filenames)
    openFileDialog.Destroy()
    return filenames


def remove_double_empty_rows(df):
    for r in range(df.shape[0] - 2, -1, -1):
        dbl_rows = df.iloc[r:r + 2, :]
        if (dbl_rows == '').all().all():
            df.drop(r + 1, inplace=True)

    return df


def is_multiwell(folder_name):
    ans = False
    pattern = re.compile(r'^[a-zA-Z]\d{1,2}$')
    match = pattern.match(folder_name)

    if not match is None:
        ans = True

    return ans


def show_estimated_time(t1, nb_of_tables):
    t2 = datetime.now()
    duration = round((t2 - t1).total_seconds())
    mess = 'Estimated time for one table: {} seconds.\n\nEstimated time for {} tables: {} minutes.\n\n' \
           'Extra time is expected for the processing of the data.' \
           ''.format(duration, nb_of_tables, round(duration * nb_of_tables / 60))

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(Mbox, 'Estimated reading time', mess, 1)
        ans = future.result()

    if ans == 2:
        sys.exit('Process terminated by user')


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


if __name__ == '__main__':
    params = {}
    run(params)

# Changelog:
# v1.00: - using wxPython for the file picker, multiple sheets stored as dictionary of DataFrames (keys = sheet names)
# v1.01: - Adding default values before def run
# v1.02: - Exception handling when measurement name contains dots: 'Std. Dev. Intensity'
# v1.10: - Adding summary tab if needed with object set count.
#        - Handles multiple summary tabs and adds counting and % values
#        - Added virtual environment activation
# v1.20: - Renaming script to "Process..."
#        - Scenario D (individual process) becomes available for multiple selected files
#        - Adding a summary table on top of individual files for scenario D
# v1.30: - User now selects the scenario in the code instead of selecting actions
#        - Adding scenario E for multiwell batch.
# v1.40: - Scenario E >> Combine summary values per well in the same column in the 'Main_Summary' table
#        - Adding a magicGui for the selection of the scenario
# v1.50: - Adding scenario F
#        - Renaming main summary as 'Analysis_Summary'
# v1.51: - Providing estimation on how long it takes to read one table if more than 10 tables are selected
# TODO: progress bar with file in Recipes folder: '_progress_bar_file 1_from 10_'
