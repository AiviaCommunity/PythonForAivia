# -------- Activate virtual environment -------------------------
import os
import ctypes
import sys
from pathlib import Path

def search_activation_path():
    for i in range(5):
        final_path = str(Path(__file__).parents[i]) + '\\env\\Scripts\\activate_this.py'
        if os.path.exists(final_path):
            return final_path
    return ''

activate_path = search_activation_path()

if os.path.exists(activate_path):
    exec(open(activate_path).read(), {'__file__': activate_path})
    print(f'Aivia virtual environment activated\nUsing python: {activate_path}')
else:
    error_mess = f'Error: {activate_path} was not found.\n\nPlease check that:\n' \
                 f'   1/ The \'FirstTimeSetup.py\' script was already run in Aivia,\n' \
                 f'   2/ The current python recipe is in one of the "\\PythonEnvForAivia\\" subfolders.'
    ctypes.windll.user32.MessageBoxW(0, error_mess, 'Error', 0)
    sys.exit(error_mess)
# ---------------------------------------------------------------

import pandas as pd
import wx
import concurrent.futures
import openpyxl.utils.cell
from magicgui import magicgui
import re
from datetime import datetime
import numpy as np
from skimage.io import imread, imsave

"""
Convert multiple Excel spreadsheets (in the same input folder) exported from Aivia into a single Excel file.
Options:
    A - Multiple files selected, spreadsheets do not contain timepoints, data is combined in the same column (stacked data)
    B - Multiple files selected, spreadsheets do not contain timepoints, data is combined in multiple columns 
        (1 column = data from 1 spreadsheet)
    C - Multiple files selected, spreadsheets do CONTAIN timepoints, data is combined in the same column
        (1 column = 1 timepoint)
    D - List of files processed as single files, spreadsheets do not contain timepoints, 
        measurement tabs are combined as multiple columns (1 column = 1 measurement, 1 tab = 1 object subset)
    E - Same as D but only one file is selected. Expecting Aivia 11.0 subfolder hierarchy from a multiwell Workflow
        batch processing 
    ... see other info above in scenario descriptions

WARNING: This currently works only under the following conditions:
    - The file were exported from Aivia as Excel files (not CSV)
    - There is no time dimension
    - The default row/column ordering was not changed at export.
    - Filenames do not contain '.' characters

The converted file will be saved with the same name as the original but with
"..._grouped" appended to the end.

Requirements
------------
pandas
openpyxl
xlrd
wxPython
magicgui

Parameters
----------
aivia_excel_file : string
    Path to the Excel file exported from Aivia.

Returns
-------
DataFrame 
    Data from the spreadsheet converted to a Pandas DataFrame.

"""

# Folder to quickly run the script on all Excel files in it
DEFAULT_FOLDER = r''

# Choice lists for the interactive form
choice_list1 = ['Excel tables are in subfolders, from a batch analysis [From Workflow / Aivia 11.0+]',
                'Excel tables are in the same folder']

choice_list2 = ['Combine all Excel tables in one big table',
                'Process Excel tables one by one']

choice_list3 = ['Data do not contain timepoints',
                'Data contains timepoints']

choice_list4 = ['Only calculate statistics (subgroup counting, %, average for some measurements)',
                'Group raw data (1 column = result from 1 image)',
                'Group raw data (1 column = 1 measurement, stack values from different images)']

# Relationship definitions (Warning: names should be the sheets in the spreadsheets)
# Example: 'Set': ['Obj1', 'Obj2']
relationships = {'Neuron Set': ['Soma Set', 'Dendrite Set', 'Dendrite Segments'],
                 'Dendrite Set': ['Dendrite Segments'],
                 'Cells': ['Cell Membranes', 'Cytoplasm', 'Nucleus', 'Vesicles - ']}

# Due to discrepancy between object name and ID header in related object meas, we can provide the correspondence below
relationship_ID_headers = {'Neuron Set': 'Neuron ID', 'Dendrite Set': 'Tree ID', 'Cells': 'Cell ID'}

# Measurements to extract, to avoid too many columns in the final table. Keywords are searched as prefix!
relationship_measurements = {'Soma Set': ['Volume '],
                             'Dendrite Set': ['Mean Diameter '],
                             'Dendrite Segments': ['Mean Diameter ', 'Total Path Length ', 'Branch Angle'],
                             'Cell Membranes': [],
                             'Cytoplasm': [],
                             'Nucleus': [],
                             'Vesicles - ': ['Is In Cytoplasm', 'Is On Cell Membrane', 'Is On Nuclear Membrane', 'Is In Nucleus']
}

# Selection of secondary relationships for which statistics (all cells) are calculated: 'Total', 'Average'.
relationships_with_stats = ['Dendrite Set', 'Dendrite Segments', 'Vesicles - ']

# Some statistics do not make any sense, so below are the ones to avoid         # TODO: filter GUI thanks to these?
relationship_measurements_stats_todrop = {'Branch Angle': 'Total',
                                          'Mean Diameter (µm)': 'Total',
                                          'Is In Cytoplasm': 'Average', 'Is On Cell Membrane': 'Average',
                                          'Is On Nuclear Membrane': 'Average', 'Is In Nucleus': 'Average'}

# reference names for the object classification results >> need to be searched in FULL tab name
class_group_ref = ['Class Group', 'Class Numb er']           # "Group" in Aivia 13.x-, "Number" in Aivia 14.x+
class_group_cut_ref = [st[-14:] for st in class_group_ref]      # Tab names are cut. In Aivia 14.x, it's 14 characters.
class_conf_ref = ['Class Confidence']
class_conf_cut_ref = [st[-14:] for st in class_conf_ref]

ui_message = "Notes for multiwell plate:" \
             "\n* If a multiwell format exists, data of images in the same well are stacked altogether" \
             "\n(becomes 1 column = 1 measurement from 1 well)." \
             "\n* For statistics, results will group images per well."

final_name_prefix = 'Analysis Summary'


@magicgui(persist=True, layout='form',
          ch1={"label": "Excel table location:\n(tooltip available)", "widget_type": "RadioButtons", 'choices': choice_list1},
          ch2={"label": "Multi-table process:", "widget_type": "RadioButtons", 'choices': choice_list2},
          ch3={"label": "Time dimension in data:", "widget_type": "RadioButtons", 'choices': choice_list3},
          ch4={"label": "Action to do on tables:", "widget_type": "RadioButtons", 'choices': choice_list4},
          spacer={"label": "  ", "widget_type": "Label"},
          text={"label": ui_message, "widget_type": "Label"},
          call_button="Run")
def get_scenario(ch1=choice_list1[0], ch2=choice_list2[0], ch3=choice_list3[0], ch4=choice_list4[0],
                 spacer='', text=''):
    """
    :param ch1:
        For batch result, select one xlsx table, an automatic search is performed to process other tables in the same batch.
    :param ch2:
    :param ch3:"widget_type": "LineEdit",
    :param ch4:
    :param text:
    :return:
    """
    pass


@get_scenario.ch3.changed.connect
def change_ch3_callback():
    if get_scenario.ch3.value == choice_list3[1]:
        pass                                            # TODO


@get_scenario.called.connect
def close_GUI_callback():
    get_scenario.close()


get_scenario.show(run=True)
choice_1 = get_scenario.ch1.value
choice_2 = get_scenario.ch2.value
choice_3 = get_scenario.ch3.value
choice_4 = get_scenario.ch4.value


# [INPUT Name:inputPath Type:string DisplayName:'Any channel']
# [OUTPUT Name:resultPath Type:string DisplayName:'Dummy to delete']
def run(params):
    input_p = params['inputPath']
    result_p = params['resultPath']
    global choice_list1, choice_list2, choice_list3, choice_list4
    global relationships, relationship_ID_headers, relationship_measurements
    global class_group_ref, class_group_cut_ref, class_conf_ref, class_conf_cut_ref

    do_multiple_files_as_cols = False  # Default action when combining multiple spreadsheets
    do_combine_meas_tabs = False  # Combining measurement tabs into one (for the same object subset)
    do_separated_processing = False  # If tables are processed separately
    do_scan_workflow_folders = False    # relative to batch analysis in Aivia 11.0+

    if choice_2 == choice_list2[1] and choice_4 == choice_list4[2]:
        do_separated_processing = True

    if choice_2 == choice_list2[0] and choice_4 == choice_list4[1]:
        do_multiple_files_as_cols = True        # data are not stacked, except if there are multiple images per well

    if choice_4 == choice_list4[2]:
        do_combine_meas_tabs = True

    if choice_1 == choice_list1[0]:
        do_scan_workflow_folders = True

    add_summary = False  # Used to know if the tab is missing from the beginning
    contains_tps = False  # If tables contain timepoints (form also asks the same, but this will check if true or not)

    # Choose files (or rely on an hard coded default folder)
    input_folder = DEFAULT_FOLDER
    if input_folder != "":
        # Preparing file list
        all_files = os.listdir(input_folder)
        indiv_path_list = [os.path.join(os.path.abspath(input_folder), f) for f in all_files
                           if (f.endswith('.xlsx') and not f.endswith('_grouped.xlsx') and not f.startswith('~')
                               and not f.startswith('._') and not f.startswith('Analysis Summary'))]

    else:
        indiv_path_list = pick_files()
        input_folder = os.path.dirname(indiv_path_list[0])

    # [From Workflow / Aivia 11.0+]: Collecting main folder
    batch_path = ''
    multiwell_mode = False
    well_ref_for_tables = []  # Expected structure = batch \ A1 \ Job 1 \ Measurements \
    if do_scan_workflow_folders:
        if os.path.basename(input_folder) != 'Measurements':
            error_msg = 'This folder ({}) is expected to be named "Measurements".' \
                        '\nSelect a table again...'.format(input_folder)
            stop_with_error_popup(error_msg)

        indiv_path_list = []

        # Detect batch folder
        expected_well_folder = str(Path(input_folder).parents[1])  # 2 levels up (A1 \ Job 1 \ Measurements \)
        if is_multiwell(os.path.basename(expected_well_folder)):
            batch_path = str(Path(expected_well_folder).parent)  # 1 level up
            multiwell_mode = True
        else:
            batch_path = expected_well_folder

        # Search files in subfolders        # TODO: exclude folders where measurements doesn't not exist
        if multiwell_mode:
            main_subfolders = [os.path.join(batch_path, wf) for wf in os.listdir(batch_path)
                               if os.path.isdir(os.path.join(batch_path, wf))]
        else:
            main_subfolders = [batch_path]

        for well_f in main_subfolders:  # well level if multiwell

            for fov_f in [os.path.join(well_f, fov) for fov in os.listdir(well_f)
                          if os.path.isdir(os.path.join(well_f, fov))]:  # FOV level

                for f in [x for x in os.listdir(os.path.join(fov_f, 'Measurements')) if (
                        x.endswith('.xlsx') and not x.endswith('_grouped.xlsx')
                        and not x.startswith('~') and not x.startswith('._')
                )]:
                    indiv_path_list.append(os.path.join(os.path.join(fov_f, 'Measurements'), f))
                    well_ref_for_tables.append(os.path.basename(well_f))

    if len(indiv_path_list) < 1:
        error_msg = 'No Excel file found in the selected folder:\n{}\nTry to select another folder'.format(input_folder)
        stop_with_error_popup(error_msg)

    # Prompt for user to see how many tables will be processed
    mess = '{} Excel files were detected.\nPress OK to continue.'.format(len(indiv_path_list)) + \
           '\nA confirmation popup message will inform you when the process is complete.'
    print(mess)  # for log
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(Mbox, 'Detected tables', mess, 1)
        ans = future.result()
    
    if ans == 2:
        sys.exit('Process terminated by user')

    # Sort files using subfolders numbers (necessary due to the absence of zero-filling, i.e. 8, 9, 10, 11, etc.)
    if do_scan_workflow_folders:
        print(Path(indiv_path_list[0]).parents[1].name)
        if re.match(r'.*\d+', str(Path(indiv_path_list[0]).parents[1].name)):
            # Collect all job folder names
            indiv_folders = [Path(fo).parents[1].name for fo in indiv_path_list]

            # Add zeros in front of number
            indiv_folders = [re.sub(r'\d+', str(re.search(r'\d+', z).group(0)).rjust(6, '0'), z) for z in indiv_folders]

            # Sort them with number
            sorted_folders_index = sorted(range(len(indiv_folders)), key=lambda tmp: indiv_folders[tmp])

            # Redefine the list of files and list of well references
            tmp_list = indiv_path_list
            indiv_path_list = [tmp_list[ind] for ind in sorted_folders_index]
            tmp_list = well_ref_for_tables
            well_ref_for_tables = [tmp_list[ind] for ind in sorted_folders_index]

    # Starting point for scenario with multiple files processed individually
    if do_separated_processing:
        final_input_list = indiv_path_list  # To prepare loop to process files independently
    else:
        final_input_list = [indiv_path_list[0]]  # Dummy entry which is 1 item to run the process once

    # OUTPUT folder
    output_folder = batch_path if do_scan_workflow_folders else os.path.abspath(input_folder)

    # Init for a summary table          TODO: link to action on table >> which action
    df_big_summary = pd.DataFrame()
    tab_ind = 0  # Tab index for pooling of summary tabs from the same well
    temp_tab = []  # Used to store tabs from the same well
    well_ref = ''

    # Evaluate time if number of file is > 10
    if len(final_input_list[:]) > 10:
        t1 = datetime.now()

    # Init measurement statistics selection for relationships which would be constant in the main loop
    relationship_meas_stats_sel = []

    # Init class names if applicable
    class_names = []

    # Main LOOP -----------------------------------------------------------------------------------------------
    for file_index, input_file in enumerate(final_input_list):
        if len(final_input_list) > 1:
            indiv_path_list = [input_file]  # List of tables is trimmed down to one item to invoke separate processing

        # Evaluate time if number of file is > 10
        if len(indiv_path_list[:]) > 10:
            t1 = datetime.now()

        print('Processing: ', input_file)

        # Reading first file to collect info
        df_raw_1 = pd.read_excel(indiv_path_list[0], sheet_name=None, engine='openpyxl')

        # Check if timepoints exist (would not allow combining tabs as different columns)
        all_tabs = list(df_raw_1.keys())
        last_tab = all_tabs[-1]
        if df_raw_1[last_tab].shape[1] > 2:  # Checking no of columns in last sheet as 1st might be summary
            contains_tps = True
            # Check if scenario is compatible
            if choice_3 == choice_list3[0]:
                error_msg = f'Timelapse data detected but option without timepoints was selected:' \
                            f'\n\n{choice_3}'
                stop_with_error_popup(error_msg)

        # Check if summary tab is present or not
        if not any('Summary' in y for y in all_tabs):
            add_summary = True

        # defining output name
        if do_multiple_files_as_cols or (do_combine_meas_tabs and do_multiple_files_as_cols):
            output_basename = final_name_prefix + '_All results.xlsx'
        elif do_combine_meas_tabs and not do_separated_processing:
            output_basename = final_name_prefix + '_All results stacked.xlsx'
        else:
            output_basename = '{}_grouped.xlsx'.format('.'.join(os.path.basename(indiv_path_list[0]).split('.')[:-1]))
        output_file = os.path.join(output_folder, output_basename.replace('.aivia', ''))

        df_grouped = {}  # init

        # Collect tab names from first file
        tab_names_ref = df_raw_1.keys()
        real_tab_names_ref = [clean_tab_name(df_raw_1[tmp_t].columns[0], ) for tmp_t in df_raw_1.keys()]
        real_tab_names_ref = change_duplicate_tab_names(real_tab_names_ref)

        # First table in final table
        for i_t, tab in enumerate(df_raw_1.keys()):
            df_grouped[real_tab_names_ref[i_t]] = df_raw_1[tab]

        if len(indiv_path_list) > 1:

            if do_multiple_files_as_cols and not contains_tps:
                do_combine_meas_tabs = False  # not possible as columns = measurements

                # Detect multiwell batch
                process_wells = is_multiwell(well_ref_for_tables[0])  # TODO: Remove summary tab if True?

                # Renaming column headers for first table
                if not process_wells:
                    prefix_name = clean_excel_name(os.path.basename(indiv_path_list[0]))
                    for t in real_tab_names_ref:
                        df_grouped[t].rename(columns={df_grouped[t].columns[-1]: prefix_name}, inplace=True)

                # Loop
                tab_names = ''
                first_well = True
                temp_df = {}
                print('-- {} processed (1/{}).'.format(os.path.basename(indiv_path_list[0]), len(indiv_path_list)))

                for f_index, f in enumerate(indiv_path_list[1:]):
                    df_raw = pd.read_excel(f, sheet_name=None, engine='openpyxl')
                    tab_names = df_raw.keys()

                    # Check if well changes
                    same_well = (well_ref_for_tables[f_index + 1] == well_ref_for_tables[f_index])

                    if tab_names == tab_names_ref:
                        prefix_name = clean_excel_name(os.path.basename(f))

                        # Start looping over the different sheets
                        for i_t, t in enumerate(tab_names):
                            real_t = real_tab_names_ref[i_t]
                            if process_wells and same_well:
                                if first_well:
                                    df_grouped[real_t] = pd.concat([df_grouped[real_t], df_raw[t]], axis=0, ignore_index=True)
                                else:
                                    # Add data as stacked in the current (temp) column
                                    temp_df[real_t] = pd.concat([temp_df[real_t], df_raw[t]], axis=0, ignore_index=True)

                            else:  # Add data as a new column
                                # Push existing temp column to final table
                                if process_wells:
                                    if not first_well:
                                        df_grouped[real_t] = pd.concat([df_grouped[real_t], temp_df[real_t].iloc[:, 1]], axis=1)

                                    # Storing data in a new stacked column
                                    temp_df[real_t] = df_raw[t]

                                else:
                                    df_grouped[real_t] = pd.concat([df_grouped[real_t], df_raw[t].iloc[:, 1]], axis=1)

                                # Renaming column header
                                if process_wells:   # header is well name
                                    df_grouped[real_t].rename(
                                        columns={df_grouped[real_t].columns[-1]: well_ref_for_tables[f_index]}, inplace=True)
                                else:               # Header is file name
                                    df_grouped[real_t].rename(columns={df_grouped[real_t].columns[-1]: prefix_name}, inplace=True)

                        if not same_well:
                            first_well = False

                        print('-- {} processed ({}/{}).'.format(os.path.basename(f), f_index + 2, len(indiv_path_list)))

                    else:
                        print('WARNING: {} table was excluded because table tabs are different from first table'.format(f))

                    # Evaluate time for the processing of one table
                    if len(indiv_path_list[:]) > 10 and f_index == 0:
                        show_estimated_time(t1, len(indiv_path_list[:]))

                # Push the latest stacked column in the final table (in case of multiwell)
                for r_t in real_tab_names_ref:
                    if process_wells:
                        if temp_df:
                            df_grouped[r_t] = pd.concat([df_grouped[r_t], temp_df[r_t].iloc[:, 1]], axis=1)
                        df_grouped[r_t].rename(columns={df_grouped[r_t].columns[-1]: well_ref_for_tables[-1]}, inplace=True)

            else:  # Stacked data from different tables TODO: stacking of timelapse data ok?
                # Adding prefix (file name) to first column
                prefix_name = clean_excel_name(os.path.basename(indiv_path_list[0]))
                for r_t in real_tab_names_ref:
                    df_grouped[r_t].iloc[:, 0] = [prefix_name + "_" + r for r in df_grouped[r_t].iloc[:, 0]]

                # Loop
                for f in indiv_path_list[1:]:
                    if indiv_path_list != output_basename:  # avoids including an existing grouped table
                        df_raw = pd.read_excel(f, sheet_name=None, engine='openpyxl')
                        tab_names = df_raw.keys()

                        if tab_names == tab_names_ref:
                            # Start looping over the different sheets
                            prefix_name = clean_excel_name(os.path.basename(f))
                            for i_t, t in enumerate(tab_names):
                                r_t = real_tab_names_ref[i_t]
                                # Adding prefix (file name) to first column in the raw table
                                df_raw[t].iloc[:, 0] = [prefix_name + "_" + r for r in df_raw[t].iloc[:, 0]]

                                # Merging to previous grouped data
                                df_grouped[r_t] = pd.concat([df_grouped[r_t], df_raw[t]], axis=0, ignore_index=True)

                        # Evaluate time for the processing of one table
                        if len(indiv_path_list[:]) > 10 and f == indiv_path_list[1]:
                            show_estimated_time(t1, len(indiv_path_list[1:]))

        # --- COMBINE TABS into one if no timepoints in data (scenario A-D...) ------------------------------
        summary_lbl = 'Summary'     # Important for the further processing of the summary tab which name can vary

        if not contains_tps and (do_combine_meas_tabs or do_multiple_files_as_cols):
            # Init
            col_headers = ['Summary', *list(df_grouped[list(df_grouped.keys())[-1]].columns[1:])]
            empty_init_list = [['']] * len(col_headers)
            empty_init_row = dict(zip(col_headers, empty_init_list))
            empty_list = [' '] * len(col_headers)
            empty_row = dict(zip(col_headers, empty_list))

            # Calculate object counts
            df_summary_to_add = pd.DataFrame(empty_init_row)
            total_counts = []
            t = 0
            grand_total = 0

            # Combine measurement tabs into one ----------------------
            if do_combine_meas_tabs:
                df_grouped = combine_tabs(df_grouped)

                # Collecting all measurements exact names
                all_meas_names = []
                for tmp_df in df_grouped.values():
                    if not tmp_df.empty:
                        if not 'summary' in str(tmp_df.columns[0]).lower():
                            all_meas_names.extend(tmp_df.columns[1:])

                # Specific to neurons: split dendrite trees from segments
                if 'Dendrite Set' in df_grouped.keys():
                    df_grouped_to_add = {}
                    order_of_keys = []
                    for k in df_grouped.keys():
                        order_of_keys.append(k)
                        if 'Dendrite Set' in k:
                            df_tmp_set, df_tmp_segm = split_dendrite_set_and_segments(df_grouped[k])
                            df_grouped[k] = df_tmp_set
                            if df_tmp_segm.shape[0] > 0:
                                segment_set_name = k.replace(' Set', ' Segments')
                                df_grouped_to_add[segment_set_name] = df_tmp_segm
                                order_of_keys.append(segment_set_name)

                    if len(df_grouped_to_add) > 0:
                        df_grouped_tmp = df_grouped.copy()     # duplicate for reordering of keys
                        df_grouped.clear()
                        for n in order_of_keys:
                            if n not in df_grouped_tmp.keys():
                                df_grouped[n] = df_grouped_to_add[n]
                            else:
                                df_grouped[n] = df_grouped_tmp[n]

                # --- Process RELATIONSHIPS between object sets (see definition before the def run) ---
                # Select all tabs where the primary object exists         # E.g. 'Cells (1)'
                relationship_parent_tabs = []
                for rel_k in relationships.keys():
                    relationship_parent_tabs += [[it_k, rel_k] for it_k in df_grouped.keys() if rel_k in it_k]

                # Process parent tab 1 by 1
                for [p_tab, p_name] in relationship_parent_tabs:    # p_tab = name of tab, p_name = name of object set???
                    p_tab_suffix = p_tab.replace(p_name, '')      # Important when multiple object sets exists (' (2)')

                    # List of available secondary objects (tabs) for the same primary object
                    secondary_tab_list = [it_s for it_s in df_grouped.keys() if is_same_object_set(it_s, p_tab_suffix)]

                    # Check validity of tabs for association with parent tab
                    # v1.60 gives the ability to provide only the beginning of the object name ('Vesicles - ')
                    # v1.61 gives the ability to search relationships for cell components even if renamed
                    # It also provides relationships of multiple secondary objects beginning with the same name
                    s_valid_tab_list = [[], []]     # 0 = tab name, 1 = object type
                    for s_tab in secondary_tab_list:
                        obj_type = [rel_s for rel_s in relationships[p_name] if rel_s in s_tab]
                        if obj_type or p_name == 'Cells':
                            id_header = relationship_ID_headers[p_name]
                            if id_header in df_grouped[s_tab].columns:
                                if not obj_type:
                                    obj_type = relationships['Cells'][-1]
                                s_valid_tab_list[0].append(s_tab)
                                s_valid_tab_list[1].append(obj_type[0])
                            else:
                                if 'Vesicle' in s_tab:
                                    Mbox('Warning', f'Relationship detected between {p_name} and {s_tab} '
                                                    f'but "Cell ID" column is missing...', 1)

                    if s_valid_tab_list:
                        # Collect all possible measurements from the first Excel table
                        available_meas = []
                        for s_t in s_valid_tab_list[0]:
                            available_meas += df_grouped[s_t].columns[1:].tolist()

                        # GUI to select measurements and statistics
                        @magicgui(persist=True, layout='horizontal',
                                  total_selection={"label": "Total:", "widget_type": "Select", 'choices': available_meas},
                                  average_selection={"label": "Average:", "widget_type": "Select", 'choices': available_meas})
                        def meas_gui_selector(total_selection=available_meas[0], average_selection=available_meas[0]):
                            pass

                        @meas_gui_selector.called.connect
                        def close_GUI_callback():
                            meas_gui_selector.close()

                        if not relationship_meas_stats_sel:
                            meas_gui_selector.show(run=True)        # returns only the selected items

                            relationship_meas_stats_sel = [
                                meas_gui_selector.total_selection.value,
                                meas_gui_selector.average_selection.value
                            ]

                        # Calculating relationship-based stats
                        id_header = relationship_ID_headers[p_name]
                        for [s_tab, obj_type] in zip(s_valid_tab_list[0], s_valid_tab_list[1]):
                            prefix = s_tab + '.'

                            for stat_type in ['Count', 'Total', 'Average']:
                                print('Collecting {} for ({}) from [{}] '
                                      'to be reported for [{}]'.format(stat_type, relationship_meas_stats_sel[0],
                                                                       s_tab, p_tab))
                                df_grouped[p_tab] = calculate_relation_stats(df_grouped[p_tab], df_grouped[s_tab],
                                                                             id_header, prefix, obj_type,
                                                                             relationship_meas_stats_sel[0], stat_type)

                    # Reset of list
                    secondary_tab_list.clear()

                # Collecting summary values
                for k in df_grouped.keys():
                    if not k.endswith('Summary'):
                        total_counts.append(df_grouped[k].shape[0])
                        grand_total += total_counts[t]
                        new_row = dict(zip(list(empty_row.keys()), ['Total number_{}'.format(k), total_counts[t]]))
                        df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Report class group counts if existing. Valid for a single classifier     TODO: expand to multiple classifiers
                        if any([cg_name in ' '.join(df_grouped[k].columns) for cg_name in class_group_ref]):
                            # Retrieving names of class tabs (group and confidence)
                            class_col_names = get_class_col_names(df_grouped[k])
                            class_data = df_grouped[k][class_col_names]
                            no_groups = int(class_data[class_col_names[0]].max())

                            # Collect names with a GUI
                            if not class_names:
                                class_names = get_class_names(no_groups)

                            if no_groups > 1:
                                group_count = [0] * no_groups
                                # Counting per class
                                for g in range(1, no_groups + 1):
                                    group_count[g - 1] = class_data[class_col_names[0]][class_data[class_col_names[0]] == g].count()
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['Total number_{}_Class {}'.format(k, class_names[g-1]), group_count[g - 1]]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)
                                # % of class
                                for g in range(1, no_groups + 1):
                                    percent = '{:.1%}'.format(group_count[g - 1] / total_counts[t])
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['{}_% of Class {}'.format(k, class_names[g-1]), percent]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)
                                # Confidence mean
                                for g in range(1, no_groups + 1):
                                    mean_conf = '{:.2}'.format(class_data.loc[class_data[class_col_names[0]] == g, class_col_names[1]].mean())
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['{}_Confidence Average for Class {}'.format(k, class_names[g-1]), mean_conf]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)
                                # Confidence std
                                for g in range(1, no_groups + 1):
                                    std_conf = '{:.2}'.format(class_data.loc[class_data[class_col_names[0]] == g, class_col_names[1]].std())
                                    new_row = dict(zip(list(empty_row.keys()),
                                                       ['{}_Confidence StDev for Class {}'.format(k, class_names[g-1]), std_conf]))
                                    df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                            # Add an empty row after each object set if classes exists
                            df_summary_to_add = df_summary_to_add.append(empty_row, ignore_index=True)

                        t += 1

            # Group multiple tables as individual columns in a master table ---------------
            if do_multiple_files_as_cols:
                # Chasing counts only for object sets, not for single measurements
                object_set_ref = ''
                for k in df_grouped.keys():
                    _, object_set = get_split_name(k)
                    if object_set == '':        # If only one object set, name is not present
                        object_set = 'Object 1'

                    if not k.endswith('Summary') and not any([k.endswith(cgref) for cgref in class_group_cut_ref]):
                        if object_set != object_set_ref:
                            total_counts.append(df_grouped[k].count()[1:])
                            new_row = dict(zip(list(empty_row.keys()),
                                               ['Total number_{}'.format(object_set), *total_counts[t]]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                            object_set_ref = object_set
                            t += 1

                    elif any([k.endswith(cgref) for cgref in class_group_cut_ref]):
                        class_tab_names = get_class_tab_names(df_grouped)
                        class_group_data = df_grouped[class_tab_names[0]]
                        group_count = [[]] * (len(df_grouped[k].columns) - 1)
                        total_counts_class = [0] * (len(df_grouped[k].columns) - 1)
                        group_max = 0

                        for col_index, col_name in enumerate(class_group_data.columns[1:]):
                            col_values = class_group_data[col_name].dropna().map(int)
                            no_groups = col_values.max()
                            if no_groups > group_max:
                                group_max = no_groups
                            group_count[col_index] = [col_values[col_values == g + 1].count() for g in range(no_groups)]
                            total_counts_class[col_index] = len(col_values)

                        # Reconstitute data per row (per group) and add values
                        group_count_per_row = [0] * len(group_count)
                        class_percent_per_row = [[''] * len(group_count) for g in range(group_max)]
                        for g in range(group_max):
                            for c in range(len(group_count)):
                                if len(group_count[c]) >= g + 1:
                                    group_count_per_row[c] = group_count[c][g]
                                    class_percent_per_row[g][c] = '{:.1%}'.format(
                                        group_count_per_row[c] / total_counts_class[c])
                                else:
                                    group_count_per_row[c] = 0
                                    class_percent_per_row[g][c] = '{:.1%}'.format(0)

                            new_row = dict(zip(list(empty_row.keys()),
                                               ['Total number_{}_Class {}'.format(object_set, g + 1),
                                                *group_count_per_row]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Add percentages
                        for g in range(group_max):
                            new_row = dict(zip(list(empty_row.keys()),
                                               ['{}_% of Class {}'.format(object_set, g + 1), *class_percent_per_row[g]]))
                            df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

                        # Add an empty row after each object set if classes exists
                        df_summary_to_add = df_summary_to_add.append(empty_row, ignore_index=True)

                        t += 1

            # Adding percentages of objects if multiple object sets exists  TODO: Check for cells that % is only for vesicles
            if do_combine_meas_tabs:  # TODO: for do_multiple_files_as_cols too??
                if len(total_counts) > 1:
                    # Collect tab names without summary
                    df_grouped_keys_nosum = [k for k in df_grouped.keys() if not k.endswith('Summary')]
                    for t in range(len(total_counts)):
                        val = '{:.1%}'.format(total_counts[t] / grand_total)
                        new_row = {'Summary': '% of {}'.format(df_grouped_keys_nosum[t]), 'Frame 0': val}
                        df_summary_to_add = df_summary_to_add.append(new_row, ignore_index=True)

            # Add the summary tab
            if add_summary:
                df_summary = pd.DataFrame(empty_row)
                df_grouped[summary_lbl] = df_summary

            else:
                # Get the name of the first summary tab / TODO: process object groups independently?
                summary_lbls = [su for su in df_grouped.keys() if su.endswith('Summary')]
                summary_lbl = summary_lbls[0]

                # Replace 'Summary' header by real header in the additional summary data df
                df_summary_to_add.rename(columns={df_summary_to_add.columns[0]: summary_lbl}, inplace=True)

            # Put zeros if some summary values are NaN
            if df_grouped[summary_lbl][df_grouped[summary_lbl].columns[1]].isnull().sum() > 0:
                df_grouped[summary_lbl][df_grouped[summary_lbl].columns[1]].fillna(0, inplace=True)

            # Combine summary values as average when same summary results are stacked
            # (gives average for a whole well for instance)
            grouped_size = df_grouped[summary_lbl].groupby(by=df_grouped[summary_lbl].columns[0],
                                                           as_index=False, sort=False).size()
            if grouped_size.iloc[:, 1].max() > 1:
                df_grouped[summary_lbl] = df_grouped[summary_lbl].groupby(by=df_grouped[summary_lbl].columns[0],
                                                                          as_index=False, sort=False).mean()
                # Put Average in front of each measurement
                df_grouped[summary_lbl][df_grouped[summary_lbl].columns[0]] = 'Average_' + df_grouped[summary_lbl][df_grouped[summary_lbl].columns[0]]

            # Merge with potential existing summary tab
            df_grouped[summary_lbl] = pd.concat([df_summary_to_add, df_grouped[summary_lbl]], axis=0, ignore_index=True)

            # Removing double empty rows
            df_grouped[summary_lbl] = remove_double_empty_rows(df_grouped[summary_lbl])

        # --- WRITING EXCEL FILES -----------------------------------------------------------------------------------
        # Writing sheets to excel
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Write Summary first
            df_grouped[summary_lbl].to_excel(writer, sheet_name=summary_lbl, index=False)

            # Resizing columns
            for c in range(0, len(df_grouped[summary_lbl].columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = df_grouped[summary_lbl].iloc[:, c].map(str).str.len().max()
                writer.sheets[summary_lbl].column_dimensions[col_letter].width = len_longest_text * 1.5

            for sh in [d for d in df_grouped.keys() if d != summary_lbl]:
                df_grouped[sh].to_excel(writer, sheet_name=sh, index=False)

                # Resizing columns
                for c in range(0, len(df_grouped[sh].columns)):
                    col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                    len_longest_text = len(str(df_grouped[sh].columns[c]))
                    if c == 0 and df_grouped[sh].shape[0] > 1:  # First column with measurement name and object names
                        if len(str(df_grouped[sh].iloc[1, 0])) > len_longest_text:
                            len_longest_text = len(str(df_grouped[sh].iloc[1, 0]))
                    if len_longest_text < 10:
                        len_longest_text = 10
                    writer.sheets[sh].column_dimensions[col_letter].width = len_longest_text

        # --- FINAL SUMMARY ----------------------------------------------------------------------------------------
        # Create a final summary file if multiple tables were saved
        if len(final_input_list) > 1 and not do_multiple_files_as_cols:
            filename = os.path.basename(indiv_path_list[0]).split('.')[0]

            if multiwell_mode:          # replacing 'do_scan_workflow_folders'
                # Checking well name compared to previous table
                current_well = well_ref_for_tables[file_index]
                if current_well != well_ref or file_index == len(final_input_list) - 1:
                    if temp_tab:
                        # Combine summary tabs and add to final super table

                        # Add latest summary tab to final super table
                        if file_index == len(final_input_list) - 1:
                            temp_tab.append(df_grouped[summary_lbl][df_grouped[summary_lbl].columns[1]])

                        if len(temp_tab) > 1:
                            # Add empty table to create space between parameters
                            temp_tab.append(pd.Series('', index=range(temp_tab[0].shape[0])))

                            # Init table
                            well_tab = temp_tab[0].iloc[[0]]

                            # Add first row of other tables
                            for tab in temp_tab[1:]:
                                well_tab = well_tab.append(pd.Series(tab.iloc[[0]]), ignore_index=True)

                            # Add other rows
                            for r in range(1, temp_tab[0].shape[0]):
                                for tab in temp_tab:
                                    well_tab = well_tab.append(pd.Series(tab.iloc[[r]]), ignore_index=True)

                            # Drop empty rows
                            if isinstance(well_tab, pd.Series):
                                well_tab = well_tab.to_frame()
                            well_tab = remove_double_empty_rows(well_tab)

                        else:
                            well_tab = temp_tab[0]

                        # Push into the big summary
                        if df_big_summary.empty:
                            df_big_summary = well_tab  # init
                            # Rename header of well column
                            df_big_summary.rename(columns={df_big_summary.columns[0]: well_ref}, inplace=True)

                        else:
                            df_big_summary[well_ref] = well_tab

                    well_ref = current_well
                    temp_tab = []

                # Add first column only if temp_tab is empty
                if not temp_tab:
                    temp_tab.append(df_grouped[summary_lbl][df_grouped[summary_lbl].columns[0]])

                # Add data to be combined
                temp_tab.append(df_grouped[summary_lbl][df_grouped[summary_lbl].columns[1]])

            else:
                if df_big_summary.empty:
                    df_big_summary = df_grouped[summary_lbl]

                    # Rename header of 2nd column
                    df_big_summary.rename(columns={df_big_summary.columns[1]: filename}, inplace=True)

                else:
                    df_big_summary[filename] = df_grouped[summary_lbl][df_grouped[summary_lbl].columns[1]]

        # --- / FINAL SUMMARY ----------------------------------------------------------------------------------------

        # Evaluate time for one table if a list of table is processed individually
        if len(final_input_list[:]) > 10 and file_index == 0:
            show_estimated_time(t1, len(final_input_list[:]))

    # Main LOOP END -------------------------------------------------------------------------------------------

    final_mess = '{} table(s) got saved here:\n{}'.format(len(final_input_list), output_folder)

    # Write final summary file if multiple tables were saved
    if len(final_input_list) > 1 and not contains_tps and not do_multiple_files_as_cols:
        # defining output name
        output_file = os.path.join(output_folder, final_name_prefix + '.xlsx')

        # Write file
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_big_summary.to_excel(writer, sheet_name=final_name_prefix, index=False)

            # Resizing columns
            for c in range(0, len(df_big_summary.columns)):
                col_letter = openpyxl.utils.cell.get_column_letter(c + 1)
                # Get longest text
                len_longest_text = max(
                    [df_big_summary.iloc[:, c].map(str).str.len().max(), len(df_big_summary.columns[c])])
                writer.sheets[final_name_prefix].column_dimensions[col_letter].width = len_longest_text

        final_mess += f'\n\nA main summary table was also saved as \'{final_name_prefix}.xlsx\'.'

    final_mess += '\n\nOutput folder explorer will now open...'

    # Message box to confirm table processing
    print(final_mess)
    Mbox('Table processed', final_mess, 0)

    # Opening the output folder in Windows
    os.startfile(output_folder)

    # Creates a zero-filled image as output
    empty_image = np.zeros_like(imread(input_p))
    if empty_image.size < 1E8:
        imsave(result_p, empty_image)
        

def get_class_col_names(df: pd.DataFrame):
    '''
    :param df: Columns of df should contain the tab names (grouped tab sheet)
    :return: List of "Class group/number" and "Class Confidence"
    '''
    class_group_match = [t_name for t_name in df.columns
                         if any([cg_ref in t_name for cg_ref in class_group_ref])]
    class_conf_match = [t_name for t_name in df.columns
                        if any([cg_ref in t_name for cg_ref in class_conf_ref])]

    # Error handling
    if len(class_group_match) == 0:
        error_mess = f'Columns for the class group/number was not found using these keywords ({class_group_ref}) among:' \
                     f'\n{df.columns}'
        stop_with_error_popup(error_mess)
    if len(class_conf_match) == 0:
        error_mess = f'Columns for the class confidence was not found using these keywords ({class_conf_ref}) among:' \
                     f'\n{df.columns}'
        stop_with_error_popup(error_mess)

    return [class_group_match[0], class_conf_match[0]]


def get_class_tab_names(dic: dict):
    '''
    :param dic: tab names (keys of dic) are expected to be truncated.
                A truncated ref of names is used here to search for a match.
    :return: List of "Class group/number" and "Class Confidence"
    '''
    class_group_match = [t_name for t_name in dic.keys()
                         if any([cg_ref in t_name for cg_ref in class_group_cut_ref])]
    class_conf_match = [t_name for t_name in dic.keys()
                        if any([cg_ref in t_name for cg_ref in class_conf_cut_ref])]

    # Error handling
    if len(class_group_match) == 0:
        error_mess = f'Excel tabs for the class group/number was not found using these keywords ({class_group_cut_ref}) ' \
                     f'among:\n{list(dic.keys())}.'
        stop_with_error_popup(error_mess)
    if len(class_conf_match) == 0:
        error_mess = f'Excel tabs for the class group/number was not found using these keywords ({class_conf_cut_ref}) ' \
                     f'among:\n{list(dic.keys())}.'
        stop_with_error_popup(error_mess)

    return [class_group_match[0], class_conf_match[0]]


# Combine tabs of a single spreadsheet file
def combine_tabs(df_raw):
    df_combined = {}
    df_temp = pd.DataFrame()
    object_name = ''
    summary_exists = False          # Used to combine multiple summary tabs if existing

    for k in df_raw.keys():
        # Don't need the summary tab if included
        if k == 'Summary' or '.Summary' in k:
            if not summary_exists:
                df_combined = {'Summary': df_raw[k]}
                summary_exists = True

                # Add header as row
                df_combined['Summary'] = pd.DataFrame([['', ''], df_combined['Summary'].columns.values.tolist()],
                                                      columns=df_combined['Summary'].columns).append(
                    df_combined['Summary'])
                # Rename 1st column name to standardize it
                df_combined['Summary'].rename(columns={df_combined['Summary'].columns[0]: 'Summary'}, inplace=True)

            else:  # If there is a 2nd summary tab (2nd object set)
                # Add header as row
                df_sum_temp = pd.DataFrame([['', ''], df_raw[k].columns.values.tolist()],
                                           columns=df_raw[k].columns).append(df_raw[k])

                df_sum_temp.columns = df_combined['Summary'].columns
                df_combined['Summary'] = pd.concat([df_combined['Summary'], df_sum_temp], axis=0, ignore_index=True)

        # First iteration with detailed objects
        elif k != 'Summary' and df_temp.empty is True:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = df_raw[k].columns[0]

            # Copying the sheet
            df_temp = df_raw[k]

            # Writing headers for the 1st and 2nd column
            if object_name == '':
                df_temp.columns = ['Objects', meas_name]
            else:
                df_temp.columns = [object_name, meas_name]

        # Fill the dataframe or create a new one if object set name changed
        else:
            # Determines what type of Aivia objects (i.e. Mesh, Slice of Cell, etc.) and measurement
            meas_name, object_name_temp = get_split_name(k)
            if meas_name == '--incomplete--':
                meas_name = (df_raw[k].columns[0])[len(object_name_temp) + 1:] \
                    if len(object_name_temp) > 0 else df_raw[k].columns[0]

            # Check if object name changed or not
            if object_name_temp != object_name and object_name_temp != '':
                # Adding prepared sheet to main series to create a new sheet
                df_combined[object_name] = df_temp

                # Now using new name as the new reference
                object_name = object_name_temp

                # Copying the current read sheet to be a new one
                df_temp = df_raw[k]

                # Writing headers for the 1st and 2nd column
                if object_name == '':
                    df_temp.columns = ['Objects', meas_name]
                else:
                    df_temp.columns = [object_name, meas_name]

            else:
                # Adding the new column to existing temp sheet
                df_temp = pd.concat([df_temp, df_raw[k].iloc[:, 1]], axis=1, join='inner')     # 2_10 Removed: , ignore_index=True

                # Adding the measurement name as a header
                df_temp.rename(columns={df_temp.columns[-1]: meas_name}, inplace=True)

    # Adding a generic name to objects if none
    if object_name == '':
        object_name = 'Object Set 1'

    # Adding last prepared sheet to main series to create a new sheet
    df_combined[object_name] = df_temp

    return df_combined


def calculate_relation_stats(df_i, df_ii, id_header, meas_prefix, obj_ii_type, measurements, stat_type):
    global relationships_with_stats

    # Measurements input needs to be a list

    df_to_add = pd.DataFrame()

    if set(df_ii.columns) & set(measurements):
        # Warning: ID of main/primary object is retrieved from its name!!!
        for obj_i_name in df_i.iloc[:, 0]:
            obj_i_id = int(obj_i_name.split(' ')[-1])            # Expecting a space before number

            # Filter rows for object II matching object I ID
            df_single_obj = df_ii[df_ii[id_header] == obj_i_id]

            # Filter columns with given measurements
            df_single_obj_filtered = df_single_obj.filter(measurements)

            # Rename columns with prefix of object II
            df_single_obj_values = df_single_obj_filtered.rename(columns=lambda x: meas_prefix + x)

            # Calculate statistics only for some object II (see definition before code)
            if obj_ii_type in relationships_with_stats:

                # No of objects II
                if stat_type == 'Count':
                    df_obj_ii_stat = pd.DataFrame({meas_prefix + 'Count': [df_single_obj_values.shape[0]]})

                # Calculate total values
                if stat_type == 'Total':
                    df_obj_ii_stat = pd.DataFrame([df_single_obj_values.sum(axis=0)])

                    # Rename columns with prefix of object II
                    df_obj_ii_stat = df_obj_ii_stat.rename(columns=lambda x: 'Total_' + x)

                # Calculate average values
                if stat_type == 'Average':
                    df_obj_ii_stat = pd.DataFrame([df_single_obj_values.mean(axis=0)])

                    # Rename columns with prefix of object II
                    df_obj_ii_stat = df_obj_ii_stat.rename(columns=lambda x: 'Average_' + x)

                # TODO: drop statistics with 'ID'?

            else:
                df_obj_ii_stat = df_single_obj_values     # if no statistics needed

            # Concatenate with df
            df_obj_ii_stat = df_obj_ii_stat.replace('nan', '')         # Replace NaN with empty strings
            df_to_add = pd.concat([df_to_add, df_obj_ii_stat], axis=0, ignore_index=True)

        # Concatenate with main object I df
        df_i = pd.concat([df_i, df_to_add], axis=1)     # 2_10 Removed: , ignore_index=True

    return df_i


def split_dendrite_set_and_segments(df):
    is_segment = df[df.columns[0]].str.match(r'^.*\sSegment')

    dendrite_set_df = df[~is_segment]
    dendrite_seg_df = df[is_segment]

    return dendrite_set_df, dendrite_seg_df


# Function to distinguish 'Cells' from 'Cells (2)' and 'Cells (3)'. Used in relationship detection.
def is_same_object_set(name, suffix):
    if suffix == '':
        ans = True if not name.endswith(')') else False
    else:
        ans = True if name.endswith(suffix) else False
    return ans


def clean_tab_name(ta_name):
    # to_change is a dict of the list of characters to be replaced
    to_change = {47: '-', 178: '2', 179: '3', 181: 'u'}

    for ord_to_change in to_change.keys():
        if any([ord(charac) == ord_to_change for charac in ta_name]):
            ta_name = ta_name.replace(chr(ord_to_change), to_change[ord_to_change])

    # Limit tab name to 30 characters because Excel can't handle more!!!
    if len(ta_name) > 30:
        ta_name = ta_name[0:28] + '..'

    return ta_name


def change_duplicate_tab_names(ta_names: list):
    for ind, t in enumerate(ta_names[1:]):
        real_ind = ind + 1
        temp_ta_names = ta_names[:real_ind].copy()   # list to compare current item to
        current_name = ta_names[real_ind]

        if any([t in temp_ta_names]):
            if current_name[-1] == '.':
                ta_names[real_ind] = current_name[:-1] + '2'
            elif current_name[-2] == '.' and current_name[-1].isnumeric():
                if int(current_name[-1]) == 9:
                    ta_names[real_ind] = current_name[:-2] + '10'
                else:
                    ta_names[real_ind] = current_name[:-1] + str(int(current_name[-1]) + 1)
            elif current_name[-2:].isnumeric():
                ta_names[real_ind] = current_name[:-2] + str(int(current_name[-2:]) + 1)
            else:
                ta_names[real_ind] = current_name[:-2] + '.2'

    return ta_names


def clean_excel_name(tmp_name: str):
    return tmp_name.removesuffix('_PrintToExcel.xlsx')


def get_split_name(txt: str):
    # Check if previous object set name is present in txt
    # prev_obj_name = '' if it is for the first measurement tab or if there is only one object set with no child objects

    if txt.startswith('Std. Dev') or '.' not in txt:
        obj_name = ''
        meas_name = txt

    else:            # Presence of an object set name expected
        obj_name = txt.split('.')[0]
        meas_name = '.'.join(txt.split('.')[1:])

        # Check if text doesn't end with '...'      # TODO: Aivia 14 changes the rule to place '...' (in the middle)
        if txt.endswith('...') or txt.endswith('..'):
            txt = txt.removesuffix('...').removesuffix('..')
            if '.' not in txt:
                obj_name = ''
            else:
                obj_name = txt.split('.')[0]
            meas_name = '--incomplete--'
            print('{}: name can\'t be retrieved from this text.'.format(txt))

    return meas_name, obj_name


def get_class_names(no_of_classes):
    @magicgui(persist=True,
              classnames={"label": f"Specify the names of the object classes, separated with only a comma\n"
                                   f"(e.g. First Class,Second Class,Unselected):\n\n"
                                   f"Number of expected classes = {no_of_classes}",
                          "widget_type": "TextEdit"},)
    def get_names(classnames=''):
        pass

    @get_names.called.connect
    def close_GUI_callback():
        get_names.close()

    get_names.show(run=True)
    collected_names = (get_names.classnames.value).split(',')

    # Check on number of classes
    final_names = [''] * no_of_classes
    if len(collected_names) > no_of_classes:
        final_names = collected_names[:no_of_classes]
        print(f'WARNING: too many class names provided. These classes were discarded: {collected_names[no_of_classes:]}')
    else:
        final_names[:len(collected_names)] = collected_names
        if len(collected_names) < no_of_classes:
            print(f'WARNING: Not enough class names were provided. {no_of_classes - len(collected_names)} names are missing.')

    return final_names


def pick_files():
    print('Starting wxPython app')
    app = wx.App()

    # Create open file dialog
    openFileDialog = wx.FileDialog(None, "Select a results table (xlsx) to process", ".\\", "",
                                   "Excel files (*.xlsx)|*.xlsx", wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)

    openFileDialog.ShowModal()
    filenames = openFileDialog.GetPaths()
    print("Selected table(s): ", filenames)
    openFileDialog.Destroy()
    return filenames


def remove_double_empty_rows(df):
    for r in range(df.shape[0] - 2, -1, -1):
        dbl_rows = df.iloc[r:r + 2, :]
        if (dbl_rows == '').all().all():
            df.drop(r + 1, inplace=True)

    return df


def is_multiwell(folder_name):
    ans = False
    pattern = re.compile(r'^[a-zA-Z]\d{1,2}$')
    match = pattern.match(folder_name)

    if not match is None:
        ans = True

    return ans


def show_estimated_time(t1, nb_of_tables):
    t2 = datetime.now()
    duration = round((t2 - t1).total_seconds())
    total_duration_tmp = round(duration * nb_of_tables / 60)
    total_duration = 1 if total_duration_tmp < 1 else total_duration_tmp
    mess = 'Estimated time for one table: {} seconds.\n\nEstimated time for {} tables: {} minutes.\n\n' \
           'Extra time is expected for the processing of the data.' \
           ''.format(duration, nb_of_tables, total_duration)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(Mbox, 'Estimated reading time', mess, 1)
        ans = future.result()

    if ans == 2:
        sys.exit('Process terminated by user')


def stop_with_error_popup(error_message):
    Mbox('Error', error_message, 0)
    sys.exit(error_message)


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


if __name__ == '__main__':
    params = {}
    run(params)

# Changelog:
# v1.00: - using wxPython for the file picker, multiple sheets stored as dictionary of DataFrames (keys = sheet names)
# v1.01: - Adding default values before def run
# v1.02: - Exception handling when measurement name contains dots: 'Std. Dev. Intensity'
# v1.10: - Adding summary tab if needed with object set count.
#        - Handles multiple summary tabs and adds counting and % values
#        - Added virtual environment activation
# v1.20: - Renaming script to "Process..."
#        - Scenario D (individual process) becomes available for multiple selected files
#        - Adding a summary table on top of individual files for scenario D
# v1.30: - User now selects the scenario in the code instead of selecting actions
#        - Adding scenario E for multiwell batch.
# v1.40: - Scenario E >> Combine summary values per well in the same column in the 'Main_Summary' table
#        - Adding a magicGui for the selection of the scenario
# v1.50: - Adding scenario F
#        - Renaming main summary as 'Analysis_Summary'
# v1.51: - Providing estimation on how long it takes to read one table if more than 10 tables are selected
# v1.52: - Starting to include relationship measurements (parent-children objects) for neurons. Stats = Total, Average.
#        - Redesigned detection of measurement names
# v1.53: - The script was not recognizing folder structure from batched workflow when it was not multiwell.
#          One subfolder level is then missing. Fixed in this version.
# v1.54: - Fixing a bug with scenario F when summary tab is not named 'Summary' exactly.
#          Also fixing a bug with formatting of 'Analysis_Summary' when not multiwell
#          Fixed wrong sorting of subfolders such as 'Job 9', 'Job 10', etc.
# v1.55: - New virtual env code for auto-activation
# v1.56: - Bug fix since Aivia 12.0 (r38705) security release for scenario F where only the summary tab is output
# v1.60: - Add Cell Analysis support for relationship grouping. Better recognition of object sets with numbers '(1)'
#        - Bug fixed at line 660 (if result table is empty)
# v1.61: - Secondary relationship sets can be named differently if the main object is "Cells"
#        - Handles NaN in summary tabs (no results detected, Aivia 12.1)
# v1.62: - Error message if time dimension is detected and scenario is not compatible
# v2.00: - New UI for more intuitive selection of options + new option for Timelapse data
#        - "Relation Count" measurement disappeared in Aivia 13.1. Counting secondary objects now done without meas.
# v2.10: - Adding selection of children measurements to be merged on the parent level (total or average)
#        - Inverted the order of calculated stats vs existing stats in existing Summary tabs
#        - Summary df now with the same header in first col
#        - Update in get_split_name function (supporting Aivia 13.1)
# v2.20: - Aivia 14.1 changed the name of columns from "Class Group" to "Class Number".
#          "Class Confidence" is not entirely written in the tab name.
# v2.21: - Tab names can be identical, causing undesired merge of data. A check ensures a different name is used.
#           ('change_duplicate_tab_names' function)
#        - Creating a zero-filled numpy output to avoid "error message" in Aivia, for small images

# TODO: progress bar with file in Recipes folder: '_progress_bar_file 1_from 10_'
# TODO: Warning message when Neuron set detected but no ID
# TODO: Group parent objects even if parent is absent from Aivia excel table (no measurement)
# TODO: Now (Aivia 11.0.1.r37805?) "Summary" tabs might take some prefix (e.g. "Meshes.Summary")
# TODO: If multiple objects and ID detected in measurement names, offer relationship option
